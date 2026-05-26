from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from src.io_utils import read_agg

OP_LABEL = {"write": "CREATE", "read": "READ", "update": "UPDATE", "delete": "DELETE"}
SCENARIO_LABEL = {"sequential": "Sequencial", "parallel": "Paralelo"}
SCALES = ["baixa", "media", "alta"]

HEADERS = [
    "database", "operation", "scenario",
    "tps_100K", "tps_1M", "tps_10M",
    "lat_p95_100K", "lat_p95_1M", "lat_p95_10M",
    "delta_tps_100K_1M_pct", "delta_tps_1M_10M_pct", "delta_tps_100K_10M_pct",
    "delta_lat_100K_1M_pct", "delta_lat_1M_10M_pct", "delta_lat_100K_10M_pct",
    "Resumo TPS 100K→10M", "Resumo Lat P95 100K→10M",
]


def _val(df, db, tb, sc, scale, col):
    r = df[
        (df["database"] == db)
        & (df["test_base"] == tb)
        & (df["simultaneity"] == sc)
        & (df["scale"] == scale)
    ]
    if r.empty:
        return None
    v = r[col].iloc[0]
    return None if pd.isna(v) else float(v)


def build_rows(df: pd.DataFrame):
    rows = []
    for db in ["mysql", "vitess"]:
        for tb, op in OP_LABEL.items():
            for sc, sc_label in SCENARIO_LABEL.items():
                rows.append({
                    "database": db,
                    "operation": op,
                    "scenario": sc_label,
                    "tps_100K": _val(df, db, tb, sc, "baixa", "tps_s_mean"),
                    "tps_1M": _val(df, db, tb, sc, "media", "tps_s_mean"),
                    "tps_10M": _val(df, db, tb, sc, "alta", "tps_s_mean"),
                    "lat_p95_100K": _val(df, db, tb, sc, "baixa", "lat_95th_mean"),
                    "lat_p95_1M": _val(df, db, tb, sc, "media", "lat_95th_mean"),
                    "lat_p95_10M": _val(df, db, tb, sc, "alta", "lat_95th_mean"),
                })
    return rows


def write_xlsx(rows, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "degradacao"

    # Cabeçalho
    for col_idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Linhas de dados + fórmulas
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["database"])
        ws.cell(row=i, column=2, value=r["operation"])
        ws.cell(row=i, column=3, value=r["scenario"])
        ws.cell(row=i, column=4, value=r["tps_100K"])
        ws.cell(row=i, column=5, value=r["tps_1M"])
        ws.cell(row=i, column=6, value=r["tps_10M"])
        ws.cell(row=i, column=7, value=r["lat_p95_100K"])
        ws.cell(row=i, column=8, value=r["lat_p95_1M"])
        ws.cell(row=i, column=9, value=r["lat_p95_10M"])

        # Fórmulas vivas — variação percentual entre escalas
        ws.cell(row=i, column=10, value=f"=IFERROR((E{i}-D{i})/D{i}*100,\"\")")
        ws.cell(row=i, column=11, value=f"=IFERROR((F{i}-E{i})/E{i}*100,\"\")")
        ws.cell(row=i, column=12, value=f"=IFERROR((F{i}-D{i})/D{i}*100,\"\")")
        ws.cell(row=i, column=13, value=f"=IFERROR((H{i}-G{i})/G{i}*100,\"\")")
        ws.cell(row=i, column=14, value=f"=IFERROR((I{i}-H{i})/H{i}*100,\"\")")
        ws.cell(row=i, column=15, value=f"=IFERROR((I{i}-G{i})/G{i}*100,\"\")")

        # Rótulos interpretativos
        # TPS: maior é melhor, então L (variação total) negativa = degradou
        ws.cell(
            row=i, column=16,
            value=(
                f'=IFERROR(IF(L{i}<0,"Degradou "&TEXT(-L{i},"0.0")&"%",'
                f'"Melhorou "&TEXT(L{i},"0.0")&"%"),"")'
            ),
        )
        # Latência: menor é melhor, então O (variação total) positiva = degradou
        ws.cell(
            row=i, column=17,
            value=(
                f'=IFERROR(IF(O{i}>0,"Degradou "&TEXT(O{i},"0.0")&"%",'
                f'"Melhorou "&TEXT(-O{i},"0.0")&"%"),"")'
            ),
        )

        # Formatação numérica
        for col in range(4, 7):  # D-F TPS
            ws.cell(row=i, column=col).number_format = "0.00"
        for col in range(7, 10):  # G-I Lat
            ws.cell(row=i, column=col).number_format = "0.00"
        for col in range(10, 16):  # J-O %
            ws.cell(row=i, column=col).number_format = "0.00"

    n_rows = len(rows)
    last_row = n_rows + 1

    # Formatação condicional — escala de cor
    # TPS (J-L): vermelho para negativo (ruim), verde para positivo (bom)
    rule_tps = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="num", mid_value=0, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )
    ws.conditional_formatting.add(f"J2:L{last_row}", rule_tps)

    # Lat P95 (M-O): verde para negativo (bom), vermelho para positivo (ruim)
    rule_lat = ColorScaleRule(
        start_type="min", start_color="63BE7B",
        mid_type="num", mid_value=0, mid_color="FFEB84",
        end_type="max", end_color="F8696B",
    )
    ws.conditional_formatting.add(f"M2:O{last_row}", rule_lat)

    # Largura das colunas
    widths = {
        "A": 10, "B": 12, "C": 12,
        "D": 12, "E": 12, "F": 12,
        "G": 13, "H": 13, "I": 13,
        "J": 14, "K": 14, "L": 14,
        "M": 14, "N": 14, "O": 14,
        "P": 26, "Q": 28,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 32

    # Congelar painel: primeira linha + 3 primeiras colunas
    ws.freeze_panes = "D2"

    # Auto filtro
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


if __name__ == "__main__":
    ROOT = Path(__file__).resolve().parent.parent
    df = read_agg(ROOT / "data" / "resultados_agg.csv")
    rows = build_rows(df)
    out_path = ROOT / "data" / "degradacao.xlsx"
    write_xlsx(rows, out_path)
    print(f"OK: salvo em {out_path}")
